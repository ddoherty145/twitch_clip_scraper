from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def create_highlights_excel(highlights_data, channels_list, filename=None, separate_sheets=True):
    """Create Excel file with channel highlights"""
    
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'highlights_output/channel_highlights_{timestamp}.xlsx'
    
    # Create workbook
    wb = Workbook()
    
    if separate_sheets:
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create separate sheet for each channel
        for channel_name in channels_list:
            ws = wb.create_sheet(title=channel_name[:31])  # Excel sheet name limit
            clips = highlights_data.get(channel_name, [])
            
            _create_highlights_sheet(ws, clips, channel_name)
        
        # Create summary sheet
        summary_ws = wb.create_sheet(title="Summary", index=0)
        _create_summary_sheet(summary_ws, highlights_data, channels_list)
    
    else:
        # Single sheet with all highlights
        ws = wb.active
        ws.title = "All Channel Highlights"
        
        # Combine all clips
        all_clips = []
        for channel_name, clips in highlights_data.items():
            all_clips.extend(clips)
        
        # Sort by view count
        all_clips.sort(key=lambda x: x.get('view_count', 0), reverse=True)
        
        _create_highlights_sheet(ws, all_clips, "All Channels")
    
    # Save file
    os.makedirs('highlights_output', exist_ok=True)
    wb.save(filename)
    return filename

def _create_highlights_sheet(ws, clips_data, sheet_title):
    """Create a worksheet with highlights data"""
    
    # Define headers
    headers = [
        "Rank", "Clip Title", "Channel", "URL", "Views", "Creator", 
        "Duration (s)", "Created At", "Game", "Thumbnail URL"
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add clip data
    for row, clip in enumerate(clips_data, 2):
        ws.cell(row=row, column=1, value=row-1)  # Rank
        ws.cell(row=row, column=2, value=clip.get('title', 'N/A'))
        ws.cell(row=row, column=3, value=clip.get('channel_name', clip.get('broadcaster_name', 'N/A')))
        ws.cell(row=row, column=4, value=clip.get('url', 'N/A'))
        ws.cell(row=row, column=5, value=clip.get('view_count', 0))
        ws.cell(row=row, column=6, value=clip.get('creator_name', 'N/A'))
        ws.cell(row=row, column=7, value=clip.get('duration', 0))
        ws.cell(row=row, column=8, value=clip.get('created_at', 'N/A'))
        ws.cell(row=row, column=9, value=clip.get('game_id', 'N/A'))
        ws.cell(row=row, column=10, value=clip.get('thumbnail_url', 'N/A'))
        
        # Alternate row coloring
        if row % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"
                )
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in range(1, len(clips_data) + 2):
            cell = ws.cell(row=row, column=col)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def _create_summary_sheet(ws, highlights_data, channels_list):
    """Create summary sheet with channel statistics"""
    
    ws.title = "Summary"
    
    # Title
    title_cell = ws.cell(row=1, column=1, value="Channel Highlights Summary")
    title_cell.font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Headers for summary table
    headers = ["Channel", "Total Highlights", "Top Views", "Average Views"]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add summary data
    for row, channel_name in enumerate(channels_list, 4):
        clips = highlights_data.get(channel_name, [])
        
        ws.cell(row=row, column=1, value=channel_name)
        ws.cell(row=row, column=2, value=len(clips))
        
        if clips:
            top_views = max(clip.get('view_count', 0) for clip in clips)
            avg_views = sum(clip.get('view_count', 0) for clip in clips) / len(clips)
            ws.cell(row=row, column=3, value=f"{top_views:,}")
            ws.cell(row=row, column=4, value=f"{avg_views:,.0f}")
        else:
            ws.cell(row=row, column=3, value="0")
            ws.cell(row=row, column=4, value="0")
        
        # Alternate row coloring
        if (row - 4) % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"
                )
    
    # Auto-adjust column widths
    for col in range(1, 5):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    # Add totals
    total_row = len(channels_list) + 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    
    total_clips = sum(len(highlights_data.get(ch, [])) for ch in channels_list)
    ws.cell(row=total_row, column=2, value=total_clips).font = Font(bold=True)
    
    # Overall top views
    all_clips = []
    for clips in highlights_data.values():
        all_clips.extend(clips)
    
    if all_clips:
        overall_top = max(clip.get('view_count', 0) for clip in all_clips)
        overall_avg = sum(clip.get('view_count', 0) for clip in all_clips) / len(all_clips)
        ws.cell(row=total_row, column=3, value=f"{overall_top:,}").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=f"{overall_avg:,.0f}").font = Font(bold=True)