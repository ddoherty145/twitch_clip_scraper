from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import os

def create_clips_excel(clips_data, filename=None):
    """Create an Excel file from the clips data."""

    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'clips_output/top_twitch_clips_{timestamp()}.xlsx'

    #Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Twitch Clips"

    #Define headers
    headers = [
        "Clip Title", "URL", "Views", "Creator", "Game/Category", "Duration (s)", "Created At", "Thumbnail URL"
    ]

    #Apply header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    #Add Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    #Add clip data
    for row, clip in enumerate(clips_data, 2):
        ws.cell(row=row, column=1, value=clip.get('title', 'N/A'))
        ws.cell(row=row, column=2, value=clip.get('url', 'N/A'))
        ws.cell(row=row, column=3, value=clip.get('view_count', 0))
        ws.cell(row=row, column=4, value=clip.get('creator_name', 'N/A'))
        ws.cell(row=row, column=5, value=clip.get('game_id', 'N/A'))
        ws.cell(row=row, column=6, value=clip.get('duration', 0))
        ws.cell(row=row, column=7, value=clip.get('created_at', 'N/A'))
        ws.cell(row=row, column=8, value=clip.get('thumbnail_url', 'N/A'))

    #Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    #Save File
    os.makedirs(clips_output, exist_ok=True)
    wb.save(filename)
    return filename
      